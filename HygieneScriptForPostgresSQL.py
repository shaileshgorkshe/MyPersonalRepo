#C:/Users/sgorkshe/AppData/Local/Programs/Python/Python312/python.exe c:/Scripts/Python/TestDB.py ericsson-dev

import psycopg2
import sys
import openpyxl.styles
from openpyxl import Workbook
from openpyxl.utils.exceptions import IllegalCharacterError
from datetime import datetime
import requests
import json

if len(sys.argv) < 2 or not sys.argv[1]:
    print("Please provide a project name as an argument.")
    sys.exit(1)

# Define Variables
url = "http://cops.onbmc.com/cops/api.php"

payload = json.dumps({
  "jsonrpc": "2.0",
  "method": "getreport",
  "params": {
    "selectors": {
      "project/name": [ "=", sys.argv[1] ]
    },
    "attributes": [
      "project/name",
      "project/parameters/DB_TYPE",
      "project/parameters/AR_DB_SERVER_NAME",
      "project/parameters/AR_DB_NAME",
      "project/parameters/AR_SERVER_DB_USERNAME",
      "project/parameters/AR_SERVER_DB_USER_PASSWORD"
    ]
  },
  "id": "getreport"
})
headers = {
  'APIKEY': '24df72c4a77c436a8195e0949fa3868a',
  'Content-Type': 'application/json'
}

response = requests.request("POST", url, headers=headers, data=payload)

# Parse the JSON response
data = json.loads(response.text)

# Check if the response contains the expected data
if 'result' in data and 'message' in data['result'] and len(data['result']['message']) > 0:
    result_message = data['result']['message'][0]

    # Extract the relevant information
    db_type = result_message['project/parameters/DB_TYPE']
    db_server = result_message['project/parameters/AR_DB_SERVER_NAME']
    db_name = result_message['project/parameters/AR_DB_NAME']
    db_username = result_message['project/parameters/AR_SERVER_DB_USERNAME']
    db_password = result_message['project/parameters/AR_SERVER_DB_USER_PASSWORD']
else:
    print("'"+sys.argv[1]+"' This does not exists in COPS.")
    sys.exit(1)


CustEnv = sys.argv[1]

#DBHost = db_server
#DBPort = "5000"
DBHost = "127.0.0.1"
DBPort = "46500"

DBType = db_type
DBname = db_name
DBUser = db_username
DBPassword = db_password
now = datetime.now()
Exec_Time = now.strftime("%Y-%m-%d %H:%M:%S")
FilePath = "C:/Scripts/Python/Result/"+CustEnv+" "+Exec_Time.replace(":","_")+".xlsx"

if DBType != "postgres" and DBType != "aurora":
    print("'"+sys.argv[1]+"' this customer's Database is not PostgresSQL OR AWS RDS Database.")
    sys.exit(1)

# PostgreSQL connection parameters
conn_params = { 'dbname': ''+DBname+'', 'user': ''+DBUser+'', 'password': ''+DBPassword+'', 'host': ''+DBHost+'', 'port': ''+DBPort+'' }

def align_cells(ws):
    # Apply formatting to the cells
    for row in ws.iter_rows(min_row=ws.min_row, max_row=ws.max_row):
        for cell in row:
            # Create a new alignment object with the desired properties
            alignment = openpyxl.styles.Alignment(horizontal="left", vertical="top", wrap_text=True)
            # Assign the new alignment object to the cell
            cell.alignment = alignment
            # Apply border styling
            border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                            right=openpyxl.styles.Side(style='thin'),
                                            top=openpyxl.styles.Side(style='thin'),
                                            bottom=openpyxl.styles.Side(style='thin'))
            cell.border = border

    # Adjust column width based on the length of the text
    for column_cells in ws.columns:
        max_length = 0
        column = openpyxl.utils.get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

def sanitize_value(value):
    # Replace illegal characters with a placeholder
    illegal_characters = ['\t', '\n', '\r']
    for char in illegal_characters:
        value = value.replace(char, ' ')
    return value

def ExeQuery(Qry, sheet_name, wb, ProblemNo, Problem, Issue, Recommendation):
    try:
        # Connect to PostgreSQL
        with psycopg2.connect(**conn_params) as conn:
            with conn.cursor() as cursor:
                cursor.execute(Qry)
                rows = cursor.fetchall()  # Fetch all rows
                if rows:
                    ws = wb.active
                    ws.append([ProblemNo, Problem, Issue, "NOT OK", Recommendation, sheet_name])
                    colnames = [desc[0] for desc in cursor.description]  # Get column headers

                    # Get or create the specified sheet
                    if sheet_name not in wb.sheetnames:
                        ws = wb.create_sheet(title=sheet_name)
                    else:
                        ws = wb[sheet_name]

                    # Write column headers
                    ws.append(colnames)

                    # Write rows
                    for row in rows:
                        sanitized_row = [sanitize_value(str(cell)) for cell in row]
                        try:
                            ws.append(sanitized_row)
                        except IllegalCharacterError as e:
                            print(f"Illegal character encountered: {e}")
                            print(sanitized_row)
                    align_cells(ws)
                else:
                    ws = wb.active
                    ws.append([ProblemNo, Problem, Issue, "OK", Recommendation, sheet_name])

        # Write data to "Query" sheet
        if "Query" not in wb.sheetnames:
            ws = wb.create_sheet("Query")
        else:
            ws = wb["Query"]
        ws.append([ProblemNo, Problem, Issue, Recommendation, Qry])
        align_cells(ws)
        ws = wb.active
        align_cells(ws)

    except psycopg2.OperationalError as e:
        # Connection failed
        ProblemNo = 1
        Problem = "Connection error"
        Recommendation = "It seems like you're encountering a connection issue with a server running on the " + DBHost + " at port " + DBPort + ". The error message ""Connection refused"" typically means that there is no server listening on the specified port, or the server is not configured to accept connections from your machine."
        ws = wb.active
        Issue = f"Connection error: {e}"  # Concatenate the strings
        ws.append([ProblemNo, Problem, Issue, "NOT OK", Recommendation])
        wb.save(FilePath)
        sys.exit(1)  # Exit the script with a non-zero status code indicating an error

if __name__ == "__main__":
    # Create a new Excel workbook
    wb = Workbook()

    ws = wb.active
    ws.append(["Customer: ", CustEnv])
    ws.append(["DB Name: ", DBname])
    ws.append(["DB Host: ", DBHost])
    ws.append(["Executed at: ", Exec_Time])
    ws.append([])
    ws.append(["Problem No.", "Problem", "Issue", "Status", "Recommendation","Refer Sheet"])

    # Check if "Query" sheet exists
    if "Query" in wb.sheetnames:
        ws = wb["Query"]
    else:
        # Create "Query" sheet if it doesn't exist
        ws = wb.create_sheet("Query")

    # Write data to "Query" sheet
    ws.append(["Problem No.", "Problem", "Issue", "Recommendation", "Query"])

    # PROBLEM#1: Check the databases having more than 10 active connections
    Threshould = 10
    ProblemNo = 1
    Problem = "Check the databases having more than " + str(Threshould) + " active connections"
    Issue = "It were found databases with the high number of active connections"
    Recommendation = "Check why Customers open so many active connections. It may be wrong configuration or unusual application pattern."

    Qry = """SELECT datname, state, client_addr, client_hostname, query
            FROM pg_stat_activity WHERE state!='idle' AND datname IN (
            SELECT datname FROM ( 	SELECT datname, COUNT(1) num_of_active_sessions
                                    FROM pg_stat_activity
                                    WHERE state!='idle' AND datname!=''
                                    GROUP BY 1
                                    HAVING COUNT(1)> """ + str(Threshould) + """
            ) M ) ORDER BY 1, 5"""
    ExeQuery(Qry, "HighNoOfActiveConnection", wb, ProblemNo, Problem, Issue, Recommendation)

    # PROBLEM#2: Check DB queries that take more than 5 minutes
    Threshould = 5
    ProblemNo = 2
    Problem = "Check DB queries that take more than " + str(Threshould) + " minutes"
    Issue = "Current Long-running queries."
    Recommendation = "Check why the query/queries take so much time. It maybe it's heavy non-optimized query. Maybe it's unusual application pattern."

    Qry = """SELECT NOW()-query_start AS runtime, pid AS process_id, datname AS db_name, client_addr, client_hostname, query
            FROM pg_stat_activity WHERE state!='idle' AND datname!='' AND pid <> pg_backend_pid() AND NOW() - query_start > '""" + str(
        Threshould) + """ minutes'::interval ORDER BY 1 DESC;"""
    ExeQuery(Qry, "CurrentLongRunningQuery", wb, ProblemNo, Problem, Issue, Recommendation)

    #PROBLEM#3: Check DB queries that run more than 1000 times per second
    Threshould = 1000
    ProblemNo = 3
    Problem = "Check DB queries that run more than "+str(Threshould)+" times per second"
    Issue = "Too frequent DB queries"
    Recommendation = "Check why the query/queries run so frequent. Maybe it's pointing to some abnormal pattern."

    Qry = """SELECT M.* FROM ( WITH
                a AS (SELECT dbid, queryid, query, calls s FROM pg_stat_statements),
                b AS (SELECT dbid, queryid, query, calls s FROM pg_stat_statements, pg_sleep(1))
            SELECT pd.datname AS db_name, a.query, SUM(b.s-a.s) AS runs_per_second
            FROM a, b, pg_database pd
            WHERE a.dbid = b.dbid AND a.queryid = b.queryid AND pd.oid=a.dbid AND pd.datname NOT IN ('postgres') AND pd.datname IS NOT NULL
            GROUP BY 1, 2 HAVING SUM(b.s-a.s) > """+str(Threshould)+""" ORDER BY 3 DESC) M;"""
    ExeQuery(Qry, "TooFrequestQuery", wb, ProblemNo, Problem, Issue, Recommendation)


    #PROBLEM#4: Which SELECT queries have read/touched the greatest number of rows?
    ProblemNo = 4
    Problem = "Check Which SELECT queries have read/touched the greatest number of rows?"
    Issue = "SELECT queries which touch the most number of rows."
    Recommendation = "These queries may benefit from adding indexes to reduce the number of rows retrieved/affected."

    Qry = """SELECT rows, query
            FROM pg_stat_statements WHERE query iLIKE '%SELECT%'
            ORDER BY rows DESC LIMIT 10;"""
    ExeQuery(Qry, "MostNoOfRows", wb, ProblemNo, Problem, Issue, Recommendation)  

    #PROBLEM#5: What queries are blocking other queries?
    ProblemNo = 5
    Problem = "What queries are blocking other queries?"
    Issue = "Queries have been blocked by any another query."
    Recommendation = "Check why the query/queries are blocking other queries. Maybe it takes significant time. Try to analyze why this DB query blocking other queries."

    Qry = """SELECT COALESCE(blockingl.relation::regclass::text, blockingl.locktype) AS locked_item,
            now() - blockinga.state_change AS waiting_duration,
            blockeda.pid AS blocked_pid,
            blockeda.query AS blocked_query,
            blockedl.mode AS blocked_mode,
            blockeda.state AS blocked_state,
            blockeda.wait_event AS blocked_wait_event,
            now()-blockeda.query_start AS blocked_runtime,
            blockinga.pid AS blocking_pid,
            blockinga.query AS blocking_query,
            blockingl.mode AS blocking_mode,
            blockinga.state AS blocking_state,
            blockinga.wait_event AS blocking_wait_event,
            now()-blockinga.query_start AS blocking_runtime
        FROM pg_locks blockedl
            JOIN pg_stat_activity blockeda ON blockedl.pid = blockeda.pid
            JOIN pg_locks blockingl ON (blockingl.transactionid = blockedl.transactionid OR blockingl.relation = blockedl.relation AND blockingl.locktype = blockedl.locktype) AND blockedl.pid <> blockingl.pid
            JOIN pg_stat_activity blockinga ON blockingl.pid = blockinga.pid AND blockinga.datid = blockeda.datid
        WHERE NOT blockedl.granted AND blockinga.datname = current_database()
        ORDER BY (now() - blockeda.query_start) DESC;"""
    ExeQuery(Qry, "BlockingQueries", wb, ProblemNo, Problem, Issue, Recommendation) 

    #PROBLEM#6: Actual connections to Max connections ratio (Threshold=5).
    Threshould = 5
    ProblemNo = 6
    Problem = "Actual connections to Max connections ratio (Threshold="+str(Threshould)+")."
    Issue = "Too high ratio of actual connections to max connections."
    Recommendation = "Check that there is enough connection slots."

    Qry = """SELECT a connection_slots_occupied, b max_connections, ROUND((a.actual_connections::float/nullif(b.max_connections::float,0))::numeric*100, 2) the_ratio
            FROM ( SELECT COUNT(1) AS actual_connections FROM pg_stat_activity ) a,
                ( SELECT setting AS max_connections FROM pg_settings WHERE name='max_connections' ) b
            WHERE ROUND((a.actual_connections::float/nullif(b.max_connections::float,0))::numeric*100, 2) > """+str(Threshould)+""";"""
    ExeQuery(Qry, "ConnectionSlots", wb, ProblemNo, Problem, Issue, Recommendation)

    #PROBLEM#7: The query/queries that allocates/allocate the connection slots (Threshold=5).
    Threshould = 5
    ProblemNo = 7
    Problem = "The query/queries that allocates/allocate the connection slots (Threshold="+ str(Threshould)+")."
    Issue = "The connection slots are occupied by the query/queries."
    Recommendation = "Check that there is enough connection slots."

    Qry = """SELECT datname, query, COUNT(1) num_of_allocated_connection_slots
            FROM pg_stat_activity 
            WHERE Query NOT IN ('COMMIT','ROLLBACK','<insufficient privilege>')
            AND datname NOT IN ('postgres') AND datname IS NOT NULL
            GROUP BY 1, 2 HAVING COUNT(1) > """+str(Threshould)+""" ORDER BY 3 DESC;"""
    ExeQuery(Qry, "ConnectionSlots", wb, ProblemNo, Problem, Issue, Recommendation)

    #PROBLEM#8: The query/queries that allocates/allocate the most connection slots (Threshold=5).
    Threshould = 5
    ProblemNo = 8
    Problem = "The query/queries that allocates/allocate the most connection slots (Threshold="+str(Threshould)+")."
    Issue = "The most of connection slots are occupied by single query."
    Recommendation = "It maybe configuration issue. It looks suspicious. because single query occupies the most connection slots of the DB instance."

    Qry = """SELECT ROUND((M.num_of_allocated_connection_slots_by_the_query::float/nullif(M.tot_allocated_slots::float,0))::numeric*100, 2) pctg, M.*
            FROM ( SELECT query, COUNT(1) num_of_allocated_connection_slots_by_the_query,
                ( SELECT COUNT(1) AS n FROM pg_stat_activity ) tot_allocated_slots
            FROM pg_stat_activity 
                WHERE Query NOT IN ('COMMIT','ROLLBACK','<insufficient privilege>')
                AND datname NOT IN ('postgres') AND datname IS NOT NULL
            GROUP BY 1, 3 HAVING COUNT(1) > """+str(Threshould)+""" ORDER BY 2 DESC ) M;"""
    ExeQuery(Qry, "ConnectionSlots", wb, ProblemNo, Problem, Issue, Recommendation)

    #PROBLEM#9: "Check the queries that occupy more than 10% of a CPU.
    Threshould = 10
    ProblemNo = 9
    Problem = "Check the queries that occupy more than "+str(Threshould)+" % of a CPU."
    Issue = "Query/queries that utilize significant portion of CPU."
    Recommendation = "Check why the query/queries take a significant portion of the CPU. Maybe it takes significant time. Maybe it's running too frequently. Try to analyze why this DB query takes a significant part of the CPU."

    Qry = """SELECT M.* FROM
            ( SELECT pss.userid, pss.dbid, pd.datname AS db_name,
                    ROUND((pss.total_exec_time + pss.total_plan_time)::numeric, 2) AS total_time, pss.calls,
                    ROUND((pss.mean_exec_time+pss.mean_plan_time)::numeric, 2) AS mean,
                    ROUND((100 * (pss.total_exec_time + pss.total_plan_time) / SUM((pss.total_exec_time + pss.total_plan_time)::numeric) OVER ())::numeric, 2) AS cpu_portion_pctg,
                    query
            FROM pg_stat_statements pss, pg_database pd
            WHERE pd.oid=pss.dbid
            ORDER BY (pss.total_exec_time + pss.total_plan_time)
            DESC LIMIT 30 ) M WHERE cpu_portion_pctg > """+str(Threshould)+""";"""
    ExeQuery(Qry, "CPUUtilizationOfQuery", wb, ProblemNo, Problem, Issue, Recommendation)

    #PROBLEM#10: Check the DB queries that take more than 5000 ms.
    Threshould = 5000
    ProblemNo = 10
    Problem = "Check the DB queries that take more than "+str(Threshould)+" ms."
    Issue = "Long-running queries."
    Recommendation = "Check why the query/queries take so much time. It may be it is a heavy non-optimized query. Maybe it's an unusual application pattern."

    Qry = """SELECT pss.userid, pss.dbid, pd.datname AS db_name,
                    ROUND((pss.total_exec_time + pss.total_plan_time)::numeric, 2) AS total_time, pss.calls,
                    ROUND((pss.mean_exec_time+pss.mean_plan_time)::numeric, 0) AS mean, query
            FROM pg_stat_statements pss, pg_database pd
            WHERE pd.oid=pss.dbid AND round((pss.mean_exec_time+pss.mean_plan_time)::numeric, 0) > """+str(Threshould)+"""
            AND Query NOT IN ('COMMIT','ROLLBACK','<insufficient privilege>')
            ORDER BY ROUND((pss.mean_exec_time+pss.mean_plan_time)::numeric, 0) DESC LIMIT 30;"""
    ExeQuery(Qry, "LongRunningQuery", wb, ProblemNo, Problem, Issue, Recommendation)

    #PROBLEM#11: Which are the unused or low-use indices?
    ProblemNo = 11
    Problem = "Which are the unused or low-use indices?"
    Issue = "Unused or low-use indices."
    Recommendation = "Check if can drop Unused or low-use indices."

    Qry = """WITH table_scans as (
            SELECT relid,
                tables.idx_scan + tables.seq_scan as all_scans,
                ( tables.n_tup_ins + tables.n_tup_upd + tables.n_tup_del ) as writes,
                        pg_relation_size(relid) as table_size
                FROM pg_stat_user_tables as tables
        ),
        all_writes as (
            SELECT sum(writes) as total_writes
            FROM table_scans
        ),
        indexes as (
            SELECT idx_stat.relid, idx_stat.indexrelid,
                idx_stat.schemaname, idx_stat.relname as tablename,
                idx_stat.indexrelname as indexname,
                idx_stat.idx_scan,
                pg_relation_size(idx_stat.indexrelid) as index_bytes,
                indexdef ~* 'USING btree' AS idx_is_btree
            FROM pg_stat_user_indexes as idx_stat
                JOIN pg_index
                    USING (indexrelid)
                JOIN pg_indexes as indexes
                    ON idx_stat.schemaname = indexes.schemaname
                        AND idx_stat.relname = indexes.tablename
                        AND idx_stat.indexrelname = indexes.indexname
            WHERE pg_index.indisunique = FALSE
        ),
        index_ratios AS (
        SELECT schemaname, tablename, indexname,
            idx_scan, all_scans,
            round(( CASE WHEN all_scans = 0 THEN 0.0::NUMERIC
                ELSE idx_scan::NUMERIC/all_scans * 100 END),2) as index_scan_pct,
            writes,
            round((CASE WHEN writes = 0 THEN idx_scan::NUMERIC ELSE idx_scan::NUMERIC/writes END),2)
                as scans_per_write,
            pg_size_pretty(index_bytes) as index_size,
            pg_size_pretty(table_size) as table_size,
            idx_is_btree, index_bytes
            FROM indexes
            JOIN table_scans
            USING (relid)
        ),
        index_groups AS (
        SELECT 'Never Used Indexes' as reason, *, 1 as grp
        FROM index_ratios
        WHERE
            idx_scan = 0
            and idx_is_btree
        UNION ALL
        SELECT 'Low Scans, High Writes' as reason, *, 2 as grp
        FROM index_ratios
        WHERE
            scans_per_write <= 1
            and index_scan_pct < 10
            and idx_scan > 0
            and writes > 100
            and idx_is_btree
        UNION ALL
        SELECT 'Seldom Used Large Indexes' as reason, *, 3 as grp
        FROM index_ratios
        WHERE
            index_scan_pct < 5
            and scans_per_write > 1
            and idx_scan > 0
            and idx_is_btree
            and index_bytes > 100000000
        UNION ALL
        SELECT 'High-Write Large Non-Btree' as reason, index_ratios.*, 4 as grp 
        FROM index_ratios, all_writes
        WHERE
            ( writes::NUMERIC / ( total_writes + 1 ) ) > 0.02
            AND NOT idx_is_btree
            AND index_bytes > 100000000
        ORDER BY grp, index_bytes DESC )
        SELECT reason, schemaname, tablename, indexname,
            index_scan_pct, scans_per_write, index_size, table_size
        FROM index_groups;"""
    ExeQuery(Qry, "UnUsedIndices", wb, ProblemNo, Problem, Issue, Recommendation) 


    #PROBLEM#12: PostgreSQL databases or tables that require a ‘VACUUM’ operation.
    ProblemNo = 12
    Problem = "PostgreSQL databases or tables that require a ‘VACUUM’ operation."
    Issue = "There are dead/live tuples which determine when a ‘VACUUM’ operation is needed."
    Recommendation = "PostgreSQL databases and tables that might require a ‘VACUUM’ operation to reclaim storage and improve performance."

    Qry = """SELECT schemaname AS schema_name, relname AS table_name, n_dead_tup AS dead_tuples,
                n_live_tup AS live_tuples, last_vacuum AS last_vacuum_time, last_autovacuum AS last_autovacuum_time,
                vacuum_count AS vacuum_frequency
            FROM pg_stat_all_tables
            WHERE n_dead_tup > 0 AND n_dead_tup > (n_live_tup * 0.1)
            ORDER BY last_vacuum_time DESC NULLS LAST, last_autovacuum_time DESC NULLS LAST;"""
    ExeQuery(Qry, "VACCUMRequired", wb, ProblemNo, Problem, Issue, Recommendation) 

    # Save the workbook to a file
    wb.save(FilePath)